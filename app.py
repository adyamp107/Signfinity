import tkinter as tk
from tkinter import filedialog, messagebox

import eel

import os
import io
import shutil
import csv
from collections import Counter
import pickle
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

import cv2
import mediapipe as mp

import base64
import re
import numpy as np
import pandas as pd

import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.metrics import classification_report, confusion_matrix

import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.models import Sequential
from tensorflow.keras.preprocessing import image

import pathlib

import warnings

from datetime import datetime

warnings.filterwarnings('ignore')

mp_drawing = mp.solutions.drawing_utils
mp_holistic = mp.solutions.holistic
holistic = mp_holistic.Holistic(min_detection_confidence=0.5, min_tracking_confidence=0.5)

root = tk.Tk()
root.withdraw()
root.wm_attributes('-topmost', 1)

cap = cv2.VideoCapture(0)

total_data = 300
count_data = 0

add_marker = False
redata_marker = False
translate_marker = False

pose_landmark = False
face_landmark = False
right_hand_landmark = False
left_hand_landmark = False

pose_bounding_box = False
face_bounding_box = False
right_hand_bounding_box = False
left_hand_bounding_box = False

on_add_page = False
on_redata_page = False
on_translate_page = False

on_add_landmark = None
on_add_image = None

on_redata_landmark = None
on_redata_image = None

count_time = 0
on_tranlsate_with_training_model = None
with_history_path = None
history_data = []
check_translate = []

landmark_data = []
image_data = []

eel.init('web')

# ============================================================================================================
# another
def stop_take_landmark_image():
    global on_add_landmark, on_add_image, on_redata_landmark, on_redata_image, landmark_data, image_data, count_data
    on_add_landmark = None
    on_add_image = None
    on_redata_landmark = None
    on_redata_image = None
    landmark_data = []
    image_data = []
    count_data = 0

# ============================================================================================================
# camera event

# ...

@eel.expose
def add_control_marker_check_event(checked):
    global add_marker
    add_marker = checked
@eel.expose
def redata_control_marker_check_event(checked):
    global redata_marker
    redata_marker = checked
@eel.expose
def setting_landmark_control_face_check_event(checked):
    global face_landmark
    face_landmark = checked
@eel.expose
def setting_landmark_control_pose_check_event(checked):
    global pose_landmark
    pose_landmark = checked
@eel.expose
def setting_landmark_control_left_hand_check_event(checked):
    global left_hand_landmark
    left_hand_landmark = checked
@eel.expose
def setting_landmark_control_right_hand_check_event(checked):
    global right_hand_landmark
    right_hand_landmark = checked
@eel.expose
def setting_bounding_box_control_face_check_event(checked):
    global face_bounding_box
    face_bounding_box = checked
@eel.expose
def setting_bounding_box_control_pose_check_event(checked):
    global pose_bounding_box
    pose_bounding_box = checked
@eel.expose
def setting_bounding_box_control_left_hand_check_event(checked):
    global left_hand_bounding_box
    left_hand_bounding_box = checked
@eel.expose
def setting_bounding_box_control_right_hand_check_event(checked):
    global right_hand_bounding_box
    right_hand_bounding_box = checked

# ============================================================================================================
# camera
@eel.expose
def camera_event():
    global add_marker, redata_marker, pose_landmark, face_landmark, right_hand_landmark, left_hand_landmark, pose_bounding_box, face_bounding_box, right_hand_bounding_box, left_hand_bounding_box, count_data, count_time, check_translate
    while(True):
        ret, frame = cap.read()
        if ret:
            frame_mark = frame
            frame_mark = cv2.cvtColor(frame_mark, cv2.COLOR_BGR2RGB)
            frame_mark.flags.writeable = False
            results = holistic.process(frame_mark)        
            frame_mark.flags.writeable = True
            frame_mark = cv2.cvtColor(frame_mark, cv2.COLOR_RGB2BGR)
            if (add_marker and on_add_page) or (redata_marker and on_redata_page) or (translate_marker and on_translate_page):
                if right_hand_landmark:
                    mp_drawing.draw_landmarks(frame_mark, results.right_hand_landmarks, mp_holistic.HAND_CONNECTIONS, 
                        mp_drawing.DrawingSpec(color=(0, 0, 255), thickness=2, circle_radius=4),
                        mp_drawing.DrawingSpec(color=(255, 0, 0), thickness=2, circle_radius=2)
                    )
                if left_hand_landmark:
                    mp_drawing.draw_landmarks(frame_mark, results.left_hand_landmarks, mp_holistic.HAND_CONNECTIONS, 
                        mp_drawing.DrawingSpec(color=(0, 0, 255), thickness=2, circle_radius=4),
                        mp_drawing.DrawingSpec(color=(255, 0, 0), thickness=2, circle_radius=2)
                    )
                if face_landmark:
                    mp_drawing.draw_landmarks(frame_mark, results.face_landmarks, mp_holistic.FACEMESH_TESSELATION, 
                        mp_drawing.DrawingSpec(color=(0, 255, 255), thickness=1, circle_radius=2),
                        mp_drawing.DrawingSpec(color=(255, 255, 0), thickness=1, circle_radius=1)
                    )    
                if pose_landmark:
                    mp_drawing.draw_landmarks(frame_mark, results.pose_landmarks, mp_holistic.POSE_CONNECTIONS, 
                        mp_drawing.DrawingSpec(color=(255, 0, 255), thickness=2, circle_radius=4),
                        mp_drawing.DrawingSpec(color=(0, 125, 255), thickness=2, circle_radius=2)
                    )
                if right_hand_bounding_box:
                    if results.right_hand_landmarks:            
                        right_hand_point = results.right_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in right_hand_point]
                        y_coordinates = [landmark.y for landmark in right_hand_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        cv2.rectangle(frame_mark, (x_min, y_min), (x_max, y_max), (0, 255, 0), 2)
                if left_hand_bounding_box:
                    if results.left_hand_landmarks:            
                        left_hand_point = results.left_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in left_hand_point]
                        y_coordinates = [landmark.y for landmark in left_hand_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        cv2.rectangle(frame_mark, (x_min, y_min), (x_max, y_max), (0, 255, 0), 2) 
                if face_bounding_box:
                    if results.face_landmarks:
                        face_point = results.face_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in face_point]
                        y_coordinates = [landmark.y for landmark in face_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30                
                        cv2.rectangle(frame_mark, (x_min, y_min), (x_max, y_max), (170, 100, 255), 2)
                if pose_bounding_box:
                    if results.pose_landmarks:
                        face_point = results.pose_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in face_point]
                        y_coordinates = [landmark.y for landmark in face_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                        cv2.rectangle(frame_mark, (x_min, y_min), (x_max, y_max), (255, 170, 100), 2)
            if on_add_landmark is not None:
                if results.right_hand_landmarks or results.left_hand_landmarks:
                    right_x_initial = [0] * 21
                    right_y_initial = [0] * 21
                    left_x_initial = [0] * 21
                    left_y_initial = [0] * 21
                    right_x_min = 0
                    right_y_min = 0
                    left_x_min = 0
                    left_y_min = 0
                    if results.right_hand_landmarks:
                        right_hand_point = results.right_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in right_hand_point]
                        y_coordinates = [landmark.y for landmark in right_hand_point]
                        right_x_min = min(x_coordinates)
                        right_y_min = min(y_coordinates)
                        x_coordinates = np.array(x_coordinates) * frame.shape[1]
                        y_coordinates = np.array(y_coordinates) * frame.shape[0]
                        right_x_initial = list(x_coordinates - right_x_min)
                        right_y_initial = list(y_coordinates - right_y_min)
                    if results.left_hand_landmarks:
                        left_hand_point = results.left_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in left_hand_point]
                        y_coordinates = [landmark.y for landmark in left_hand_point]
                        left_x_min = min(x_coordinates)
                        left_y_min = min(y_coordinates)
                        x_coordinates = np.array(x_coordinates) * frame.shape[1]
                        y_coordinates = np.array(y_coordinates) * frame.shape[0]
                        left_x_initial = list(x_coordinates - left_x_min)
                        left_y_initial = list(x_coordinates - left_y_min)  
                    new_word = on_add_landmark['new_word'].title()
                    landmark_row = [new_word] + right_x_initial + right_y_initial + left_x_initial + left_y_initial                
                    landmark_data.append(landmark_row)
                    count_data += 1
                    percentage = (count_data / total_data) * 100
                    eel.update_add_progress(percentage, count_data)()
                    if count_data >= total_data:
                        path = on_add_landmark['dataset_path']
                        csv_file = open(path, 'a', newline='')
                        csv_writer = csv.writer(csv_file)
                        for row in landmark_data:
                            csv_writer.writerow(row)
                        csv_file.close()
                        eel.to_dataset_page()()
                        stop_take_landmark_image()
            elif on_add_image is not None:
                if results.right_hand_landmarks or results.left_hand_landmarks:
                    right_x_initial = [0] * 21
                    right_y_initial = [0] * 21
                    left_x_initial = [0] * 21
                    left_y_initial = [0] * 21
                    right_x_min = 0
                    right_y_min = 0
                    left_x_min = 0
                    left_y_min = 0
                    right_cropped_frame = None
                    left_cropped_frame = None
                    cropped_frame = None
                    if results.right_hand_landmarks:
                        right_hand_point = results.right_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in right_hand_point]
                        y_coordinates = [landmark.y for landmark in right_hand_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        right_cropped_frame = frame[y_min:y_max, x_min:x_max]                    
                        if isinstance(right_cropped_frame, np.ndarray) and right_cropped_frame.size:
                            right_cropped_frame = cv2.resize(right_cropped_frame, (225, 225))
                            _, right_buffer = cv2.imencode('.jpg', right_cropped_frame)
                            right_encoded_image = np.array(right_buffer)
                            right_cropped_frame = cv2.imdecode(right_encoded_image, cv2.IMREAD_COLOR)
                    if results.left_hand_landmarks:
                        left_hand_point = results.left_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in left_hand_point]
                        y_coordinates = [landmark.y for landmark in left_hand_point]
                        x_min = int(min(x_coordinates) * frame.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        left_cropped_frame = frame[y_min:y_max, x_min:x_max]
                        if isinstance(left_cropped_frame, np.ndarray) and left_cropped_frame.size:
                            left_cropped_frame = cv2.resize(left_cropped_frame, (225, 225))                
                            _, left_buffer = cv2.imencode('.jpg', left_cropped_frame)
                            left_encoded_image = np.array(left_buffer)
                            left_cropped_frame = cv2.imdecode(left_encoded_image, cv2.IMREAD_COLOR)
                    if left_cropped_frame is not None and right_cropped_frame is not None:
                        if left_cropped_frame.ndim == right_cropped_frame.ndim and left_cropped_frame.shape[0] == right_cropped_frame.shape[0] and left_cropped_frame.dtype == right_cropped_frame.dtype:
                            cropped_frame = cv2.hconcat([left_cropped_frame, right_cropped_frame])
                            if cropped_frame.size:
                                cropped_frame = cv2.resize(cropped_frame, (225, 225))
                    elif left_cropped_frame is not None:
                        cropped_frame = left_cropped_frame
                    elif right_cropped_frame is not None:
                        cropped_frame = right_cropped_frame
                    if cropped_frame is not None and cropped_frame.size is not 0:
                        image_data.append(cropped_frame)
                        count_data += 1
                        percentage = (count_data / total_data) * 100
                        eel.update_add_progress(percentage, count_data)()
                        if count_data >= total_data:
                            path = on_add_image['dataset_path']
                            new_word = on_add_image['new_word'].title()
                            os.mkdir(f'{path}/{new_word}')
                            for index, write_image in enumerate(image_data):
                                cv2.imwrite(f'{path}/{new_word}/{new_word}_{index+1}.jpg', write_image)
                            eel.to_dataset_page()()
                            stop_take_landmark_image()
            elif on_redata_landmark is not None:
                if results.right_hand_landmarks or results.left_hand_landmarks:
                    right_x_initial = [0] * 21
                    right_y_initial = [0] * 21
                    left_x_initial = [0] * 21
                    left_y_initial = [0] * 21
                    right_x_min = 0
                    right_y_min = 0
                    left_x_min = 0
                    left_y_min = 0
                    if results.right_hand_landmarks:
                        right_hand_point = results.right_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in right_hand_point]
                        y_coordinates = [landmark.y for landmark in right_hand_point]
                        right_x_min = min(x_coordinates)
                        right_y_min = min(y_coordinates)
                        x_coordinates = np.array(x_coordinates) * frame.shape[1]
                        y_coordinates = np.array(y_coordinates) * frame.shape[0]
                        right_x_initial = list(x_coordinates - right_x_min)
                        right_y_initial = list(y_coordinates - right_y_min)
                    if results.left_hand_landmarks:
                        left_hand_point = results.left_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in left_hand_point]
                        y_coordinates = [landmark.y for landmark in left_hand_point]
                        left_x_min = min(x_coordinates)
                        left_y_min = min(y_coordinates)
                        x_coordinates = np.array(x_coordinates) * frame.shape[1]
                        y_coordinates = np.array(y_coordinates) * frame.shape[0]
                        left_x_initial = list(x_coordinates - left_x_min)
                        left_y_initial = list(x_coordinates - left_y_min)  
                    new_word = on_redata_landmark['new_word'].title()
                    landmark_row = [new_word] + right_x_initial + right_y_initial + left_x_initial + left_y_initial                
                    landmark_data.append(landmark_row)
                    count_data += 1
                    percentage = (count_data / total_data) * 100
                    eel.update_redata_progress(percentage, count_data)()
                    if count_data >= total_data:
                        path = on_redata_landmark['dataset_path']
                        df = pd.read_csv(path)
                        df = df[df['class'] != new_word]
                        df.to_csv(path, index=False)
                        csv_file = open(path, 'a', newline='')
                        csv_writer = csv.writer(csv_file)
                        for row in landmark_data:
                            csv_writer.writerow(row)
                        csv_file.close()
                        eel.to_dataset_page()()
                        stop_take_landmark_image()
            elif on_redata_image is not None:
                if results.right_hand_landmarks or results.left_hand_landmarks:
                    right_x_initial = [0] * 21
                    right_y_initial = [0] * 21
                    left_x_initial = [0] * 21
                    left_y_initial = [0] * 21
                    right_x_min = 0
                    right_y_min = 0
                    left_x_min = 0
                    left_y_min = 0
                    right_cropped_frame = None
                    left_cropped_frame = None
                    cropped_frame = None
                    if results.right_hand_landmarks:
                        right_hand_point = results.right_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in right_hand_point]
                        y_coordinates = [landmark.y for landmark in right_hand_point]
                        x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        right_cropped_frame = frame[y_min:y_max, x_min:x_max]                    
                        if isinstance(right_cropped_frame, np.ndarray) and right_cropped_frame.size:
                            right_cropped_frame = cv2.resize(right_cropped_frame, (225, 225))
                            _, right_buffer = cv2.imencode('.jpg', right_cropped_frame)
                            right_encoded_image = np.array(right_buffer)
                            right_cropped_frame = cv2.imdecode(right_encoded_image, cv2.IMREAD_COLOR)
                    if results.left_hand_landmarks:
                        left_hand_point = results.left_hand_landmarks.landmark
                        x_coordinates = [landmark.x for landmark in left_hand_point]
                        y_coordinates = [landmark.y for landmark in left_hand_point]
                        x_min = int(min(x_coordinates) * frame.shape[1]) - 30
                        y_min = int(min(y_coordinates) * frame.shape[0]) - 30
                        x_max = int(max(x_coordinates) * frame.shape[1]) + 30
                        y_max = int(max(y_coordinates) * frame.shape[0]) + 30
                        width_side = x_max - x_min
                        height_side = y_max - y_min
                        if width_side < height_side:
                            difference = height_side - width_side
                            addition = int(difference / 2)
                            x_min -= addition
                            x_max += addition
                        elif width_side > height_side:
                            difference = width_side - height_side
                            addition = int(difference / 2)
                            y_min -= addition
                            y_max += addition
                        left_cropped_frame = frame[y_min:y_max, x_min:x_max]
                        if isinstance(left_cropped_frame, np.ndarray) and left_cropped_frame.size:
                            left_cropped_frame = cv2.resize(left_cropped_frame, (225, 225))                
                            _, left_buffer = cv2.imencode('.jpg', left_cropped_frame)
                            left_encoded_image = np.array(left_buffer)
                            left_cropped_frame = cv2.imdecode(left_encoded_image, cv2.IMREAD_COLOR)
                    if left_cropped_frame is not None and right_cropped_frame is not None:
                        if left_cropped_frame.ndim == right_cropped_frame.ndim and left_cropped_frame.shape[0] == right_cropped_frame.shape[0] and left_cropped_frame.dtype == right_cropped_frame.dtype:
                            cropped_frame = cv2.hconcat([left_cropped_frame, right_cropped_frame])
                            if cropped_frame.size:
                                cropped_frame = cv2.resize(cropped_frame, (225, 225))
                    elif left_cropped_frame is not None:
                        cropped_frame = left_cropped_frame
                    elif right_cropped_frame is not None:
                        cropped_frame = right_cropped_frame
                    if cropped_frame is not None and cropped_frame.size is not 0:
                        image_data.append(cropped_frame)
                        count_data += 1
                        percentage = (count_data / total_data) * 100
                        eel.update_redata_progress(percentage, count_data)()
                        if count_data >= total_data:
                            path = on_redata_image['dataset_path']
                            new_word = on_redata_image['new_word'].title()  
                            shutil.rmtree(f'{path}/{new_word}')                            
                            os.mkdir(f'{path}/{new_word}')
                            for index, write_image in enumerate(image_data):
                                cv2.imwrite(f'{path}/{new_word}/{new_word}_{index+1}.jpg', write_image)
                            eel.to_dataset_page()()
                            stop_take_landmark_image()
            elif on_tranlsate_with_training_model is not None:
                if on_tranlsate_with_training_model['algorithm'] == 'Convolutional Neural Network':
                    cv2.rectangle(frame_mark, (0, 0), (300, 80), (255, 255, 255), -1)
                    cv2.putText(frame_mark, 'Class:', (100, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                    cv2.putText(frame_mark, 'Prob:', (20, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                    if results.right_hand_landmarks or results.left_hand_landmarks:
                        right_x_initial = [0] * 21
                        right_y_initial = [0] * 21
                        left_x_initial = [0] * 21
                        left_y_initial = [0] * 21
                        right_x_min = 0
                        right_y_min = 0
                        left_x_min = 0
                        left_y_min = 0
                        right_cropped_frame = None
                        left_cropped_frame = None
                        cropped_frame = None
                        if results.right_hand_landmarks:
                            right_hand_point = results.right_hand_landmarks.landmark
                            x_coordinates = [landmark.x for landmark in right_hand_point]
                            y_coordinates = [landmark.y for landmark in right_hand_point]
                            x_min = int(min(x_coordinates) * frame_mark.shape[1]) - 30
                            y_min = int(min(y_coordinates) * frame_mark.shape[0]) - 30
                            x_max = int(max(x_coordinates) * frame_mark.shape[1]) + 30
                            y_max = int(max(y_coordinates) * frame_mark.shape[0]) + 30
                            width_side = x_max - x_min
                            height_side = y_max - y_min
                            if width_side < height_side:
                                difference = height_side - width_side
                                addition = int(difference / 2)
                                x_min -= addition
                                x_max += addition
                            elif width_side > height_side:
                                difference = width_side - height_side
                                addition = int(difference / 2)
                                y_min -= addition
                                y_max += addition
                            right_cropped_frame = frame[y_min:y_max, x_min:x_max]                    
                            if isinstance(right_cropped_frame, np.ndarray) and right_cropped_frame.size:
                                right_cropped_frame = cv2.resize(right_cropped_frame, (225, 225))
                                _, right_buffer = cv2.imencode('.jpg', right_cropped_frame)
                                right_encoded_image = np.array(right_buffer)
                                right_cropped_frame = cv2.imdecode(right_encoded_image, cv2.IMREAD_COLOR)
                        if results.left_hand_landmarks:
                            left_hand_point = results.left_hand_landmarks.landmark
                            x_coordinates = [landmark.x for landmark in left_hand_point]
                            y_coordinates = [landmark.y for landmark in left_hand_point]
                            x_min = int(min(x_coordinates) * frame.shape[1]) - 30
                            y_min = int(min(y_coordinates) * frame.shape[0]) - 30
                            x_max = int(max(x_coordinates) * frame.shape[1]) + 30
                            y_max = int(max(y_coordinates) * frame.shape[0]) + 30
                            width_side = x_max - x_min
                            height_side = y_max - y_min
                            if width_side < height_side:
                                difference = height_side - width_side
                                addition = int(difference / 2)
                                x_min -= addition
                                x_max += addition
                            elif width_side > height_side:
                                difference = width_side - height_side
                                addition = int(difference / 2)
                                y_min -= addition
                                y_max += addition
                            left_cropped_frame = frame[y_min:y_max, x_min:x_max]
                            if isinstance(left_cropped_frame, np.ndarray) and left_cropped_frame.size:
                                left_cropped_frame = cv2.resize(left_cropped_frame, (225, 225))                
                                _, left_buffer = cv2.imencode('.jpg', left_cropped_frame)
                                left_encoded_image = np.array(left_buffer)
                                left_cropped_frame = cv2.imdecode(left_encoded_image, cv2.IMREAD_COLOR)
                        if left_cropped_frame is not None and right_cropped_frame is not None:
                            if left_cropped_frame.ndim == right_cropped_frame.ndim and left_cropped_frame.shape[0] == right_cropped_frame.shape[0] and left_cropped_frame.dtype == right_cropped_frame.dtype:
                                cropped_frame = cv2.hconcat([left_cropped_frame, right_cropped_frame])
                                if cropped_frame.size:
                                    cropped_frame = cv2.resize(cropped_frame, (225, 225))
                        elif left_cropped_frame is not None:
                            cropped_frame = left_cropped_frame
                        elif right_cropped_frame is not None:
                            cropped_frame = right_cropped_frame
                        if cropped_frame is not None and cropped_frame.size is not 0:
                            cropped_frame = cv2.cvtColor(cropped_frame, cv2.COLOR_BGR2GRAY)
                            cropped_frame = cv2.resize(cropped_frame, (48, 48))                            
                            image_array = image.img_to_array(cropped_frame)
                            image_array = np.expand_dims(cropped_frame, axis=0)
                            image_array = image_array / 255.0
                            predictions_class_index = on_tranlsate_with_training_model['trained_model'].predict(image_array)
                            predictions_probability = tf.nn.softmax(predictions_class_index[0])
                            word = on_tranlsate_with_training_model['class_names'][np.argmax(predictions_class_index)]
                            percentage = round((100 * np.max(predictions_probability)), 2)
                            cv2.putText(frame_mark, word, (100, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                            cv2.putText(frame_mark, f'{percentage}%', (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                            check_translate.append(word)
                            if len(check_translate) == 15:
                                count_words = Counter(check_translate)
                                modus_words = [word for word, max_words in count_words.items() if max_words == max(count_words.values())]
                                text_output = eel.translate_control_result_text_event()()
                                get_word = text_output.split(' ')
                                get_word = list(filter(None, text_output.split(' ')))
                                if len(get_word) == 1:
                                    if not modus_words[0] == get_word[-1]:
                                        text_output = text_output + modus_words[0] + ' '
                                        eel.send_text_translate_control_result_text_event(text_output)()
                                        count_time = 0                                        
                                        eel.translate_control_time_limit_progressbar_event(count_time)()
                                elif len(get_word) > 1:
                                    if not modus_words[0] == get_word[-1] and not modus_words[0] == get_word[-2] + ' ' + get_word[-1]:
                                        text_output = text_output + modus_words[0] + ' '                                        
                                        eel.send_text_translate_control_result_text_event(text_output)()
                                        count_time = 0                                        
                                        eel.translate_control_time_limit_progressbar_event(count_time)()
                                else:
                                    text_output = modus_words[0] + ' '
                                    eel.send_text_translate_control_result_text_event(text_output)()
                                check_translate = []  
                        else:
                            cv2.putText(frame_mark, 'Unknown', (100, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                            cv2.putText(frame_mark, '-', (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA) 
                    else:
                        cv2.putText(frame_mark, 'Unknown', (100, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                        cv2.putText(frame_mark, '-', (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA) 
                    if not eel.translate_control_result_text_event()() == '':
                        count_time += 0.5
                        eel.translate_control_time_limit_progressbar_event(count_time)()
                        if count_time == 100:
                            get_text = eel.translate_control_result_text_event()()
                            get_history_data = {
                                'date': datetime.now().strftime('%Y-%m-%d'),
                                'time': datetime.now().strftime('%H:%M:%S'),
                                'translate': get_text
                            }
                            history_data.append(get_history_data)
                            eel.send_translate_translation_list_sub_frame_event(get_history_data)()
                            eel.send_text_translate_control_result_text_event('')()
                            count_time = 0                                        
                            eel.translate_control_time_limit_progressbar_event(count_time)()
                            check_translate = []  
                    else:
                        count_time = 0
                else:
                    cv2.rectangle(frame_mark, (0, 0), (300, 80), (255, 255, 255), -1)
                    cv2.putText(frame_mark, 'Class:', (100, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                    cv2.putText(frame_mark, 'Prob:', (20, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                    if results.right_hand_landmarks or results.left_hand_landmarks:
                        right_x_initial = [0] * 21
                        right_y_initial = [0] * 21
                        left_x_initial = [0] * 21
                        left_y_initial = [0] * 21
                        right_x_min = 0
                        right_y_min = 0
                        left_x_min = 0
                        left_y_min = 0
                        if results.right_hand_landmarks:
                            right_hand_point = results.right_hand_landmarks.landmark
                            x_coordinates = [landmark.x for landmark in right_hand_point]
                            y_coordinates = [landmark.y for landmark in right_hand_point]
                            right_x_min = min(x_coordinates)
                            right_y_min = min(y_coordinates)
                            x_coordinates = np.array(x_coordinates) * frame.shape[1]
                            y_coordinates = np.array(y_coordinates) * frame.shape[0]
                            right_x_initial = list(x_coordinates - right_x_min)
                            right_y_initial = list(y_coordinates - right_y_min)
                        if results.left_hand_landmarks:
                            left_hand_point = results.left_hand_landmarks.landmark
                            x_coordinates = [landmark.x for landmark in left_hand_point]
                            y_coordinates = [landmark.y for landmark in left_hand_point]
                            left_x_min = min(x_coordinates)
                            left_y_min = min(y_coordinates)
                            x_coordinates = np.array(x_coordinates) * frame.shape[1]
                            y_coordinates = np.array(y_coordinates) * frame.shape[0]
                            left_x_initial = list(x_coordinates - left_x_min)
                            left_y_initial = list(x_coordinates - left_y_min)  
                        landmark_row = right_x_initial + right_y_initial + left_x_initial + left_y_initial              
                        X_predict = pd.DataFrame([landmark_row])
                        predictions_class = on_tranlsate_with_training_model['trained_model'].predict(X_predict)[0]
                        predictions_probability = on_tranlsate_with_training_model['trained_model'].predict_proba(X_predict)[0]
                        word = predictions_class
                        percentage = round((predictions_probability[np.argmax(predictions_probability)] * 100), 2)
                        cv2.putText(frame_mark, word, (100, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                        cv2.putText(frame_mark, f'{percentage}%', (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                        check_translate.append(word)
                        if len(check_translate) == 15:
                            count_words = Counter(check_translate)
                            modus_words = [word for word, max_words in count_words.items() if max_words == max(count_words.values())]
                            text_output = eel.translate_control_result_text_event()()
                            get_word = text_output.split(' ')
                            get_word = list(filter(None, text_output.split(' ')))
                            if len(get_word) == 1:
                                if not modus_words[0] == get_word[-1]:
                                    text_output = text_output + modus_words[0] + ' '
                                    eel.send_text_translate_control_result_text_event(text_output)()
                                    count_time = 0                                        
                                    eel.translate_control_time_limit_progressbar_event(count_time)()
                            elif len(get_word) > 1:
                                if not modus_words[0] == get_word[-1] and not modus_words[0] == get_word[-2] + ' ' + get_word[-1]:
                                    text_output = text_output + modus_words[0] + ' '                                        
                                    eel.send_text_translate_control_result_text_event(text_output)()
                                    count_time = 0                                        
                                    eel.translate_control_time_limit_progressbar_event(count_time)()
                            else:
                                text_output = modus_words[0] + ' '
                                eel.send_text_translate_control_result_text_event(text_output)()
                            check_translate = []
                    else:
                        cv2.putText(frame_mark, 'Unknown', (100, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                        cv2.putText(frame_mark, '-', (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1, cv2.LINE_AA)
                    if not eel.translate_control_result_text_event()() == '':
                        count_time += 0.5
                        eel.translate_control_time_limit_progressbar_event(count_time)()
                        if count_time == 100:
                            get_text = eel.translate_control_result_text_event()()
                            get_history_data = {
                                'date': datetime.now().strftime('%Y-%m-%d'),
                                'time': datetime.now().strftime('%H:%M:%S'),
                                'translate': get_text
                            }
                            history_data.append(get_history_data)
                            eel.send_translate_translation_list_sub_frame_event(get_history_data)()
                            eel.send_text_translate_control_result_text_event('')()
                            count_time = 0                                        
                            eel.translate_control_time_limit_progressbar_event(count_time)()
                            check_translate = []
                    else:
                        count_time = 0
            ret, buffer = cv2.imencode('.jpg', frame_mark)
            frame_bytes = buffer.tobytes()
            frame_base64 = base64.b64encode(frame_bytes).decode('utf-8')
            eel.update_image(frame_base64)()

# ============================================================================================================
# navigator
@eel.expose
def check_navigator():
    global on_add_page, on_redata_page, on_translate_page
    if on_tranlsate_with_training_model is not None:
        messagebox.showinfo('Warning', 'Sorry, you can\'t move to another page if you haven\'t stopped the translation process')
        return False
    if on_add_landmark is not None:
        messagebox.showinfo('Warning', 'Sorry, you cannot move to another page before completing the process of add landmark!') 
        return False
    if on_add_image is not None:
        messagebox.showinfo('Warning', 'Sorry, you cannot move to another page before completing the process of add image!') 
        return False
    if on_redata_landmark is not None:
        messagebox.showinfo('Warning', 'Sorry, you cannot move to another page before completing the process of redata landmark!') 
        return False
    if on_redata_image is not None:
        messagebox.showinfo('Warning', 'Sorry, you cannot move to another page before completing the process of redata image!') 
        return False
    on_add_page = False
    on_redata_page = False
    on_translate_page = False
    return True

# ============================================================================================================
# dataset

@eel.expose
def dataset_control_select_path_button_event(dataset_type):
    if dataset_type == 'Landmark File':
        dataset_path = filedialog.askopenfilename()
    elif dataset_type == 'Image Folder':
        dataset_path = filedialog.askdirectory()
    return dataset_path

@eel.expose
def dataset_control_create_button_event(dataset_path, dataset_name, dataset_type):
    if dataset_path:
        if dataset_name:
            if dataset_type == 'Landmark File':
                header = ['class']                
                x_right = []
                y_right = []
                x_left = []
                y_left = []
                for i in range(21):
                    x_right.append(f'rx{i}')
                    y_right.append(f'ry{i}')
                    x_left.append(f'lx{i}')
                    y_left.append(f'ly{i}')
                header.extend(x_right + y_right + x_left + y_left)
                dataset_name = dataset_name.split('.')[0] + '.csv'
                dataset_file = dataset_path + '/' + dataset_name
                csv_file = open(dataset_file, 'w', newline='')
                csv_writer = csv.writer(csv_file)
                csv_writer.writerow(header)
                csv_file.close()
                return dataset_file
            elif dataset_type == 'Image Folder':
                dataset_name = dataset_name.split('.')[0]
                dataset_folder = dataset_path + '/' + dataset_name
                os.makedirs(dataset_folder)
                return dataset_folder
        else:
            messagebox.showinfo('Warning', 'Please complete the dataset name section of the form!')
    else:
        messagebox.showinfo('Warning', 'Please complete the dataset path section of the form!')
    return None

@eel.expose
def dataset_control_delete_button_event(dataset_path):
    if dataset_path:
        result = messagebox.askyesno('Confirmation', f'Are you sure you want to delete this {dataset_path} dataset?')
        if result:
            if os.path.isfile(dataset_path):
                os.remove(dataset_path)
                return True
            elif os.path.isdir(dataset_path):
                shutil.rmtree(dataset_path)
                return True
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return False

@eel.expose
def dataset_selected_path_file_folder_text_event(dataset_path, dataset_type):
    if dataset_path:
        if dataset_type == 'Landmark File':
            if os.path.splitext(dataset_path)[1].lower() == '.csv':
                header = ['class']
                x_right = []
                y_right = []
                x_left = []
                y_left = []
                for i in range(21):
                    x_right.append(f'rx{i}')
                    y_right.append(f'ry{i}')
                    x_left.append(f'lx{i}')
                    y_left.append(f'ly{i}')
                header.extend(x_right + y_right + x_left + y_left)
                csv_file = open(dataset_path, 'r', newline='')
                csv_reader = csv.reader(csv_file)
                first_row = next(csv_reader, 0)
                if first_row == header:
                    labels = []
                    exit_loop = False
                    for row in csv_reader:
                        if exit_loop:
                            break
                        if not len(row) == len(header):
                            messagebox.showinfo('Warning', f'Sorry, in the {dataset_path} file there are a number of columns that do not match our dataset reader system!')
                            break
                        for column in row[1:]:
                            try:
                                float(column)
                            except ValueError:
                                messagebox.showinfo('Warning', f'Sorry, in the {dataset_path} file was found landmark data that does not comply with the requirements of our dataset reader system!')
                                exit_loop = True
                                break
                        labels.append(row[0])
                    check_labels = True                    
                    if len(labels) > 0:
                        counts = Counter(labels)
                        if not all(value == total_data for value in counts.values()):
                            messagebox.showinfo('Warning', 'Sorry, the number of landmarks in the dataset does not match!')
                            check_labels = False
                    if check_labels:
                        key = {}.fromkeys(labels)
                        unique = list(key.keys())
                        return {
                            'dataset_path': dataset_path,
                            'words': unique
                        }
                else:
                    messagebox.showinfo('Warning', f'Sorry, the header in the {dataset_path} file does not match our dataset reader system!')
                csv_file.close()
            else:
                messagebox.showinfo('Warning', 'Sorry, the dataset file is not compatible with our reader system!')
        elif dataset_type == 'Image Folder':
            dataset_labels = os.listdir(dataset_path)
            labels = []
            check_labels = True
            exit_loop = False
            for label_folder in dataset_labels:
                label = label_folder
                label_folder = dataset_path + '/' + label_folder
                if exit_loop:
                    check_labels = False
                    break
                if os.path.isdir(label_folder):
                    image_files = os.listdir(label_folder)
                    if not len(image_files) == total_data:
                        messagebox.showinfo('Warning', 'Sorry, the number of images in the dataset does not match!')
                        check_labels = False
                        break
                    else:
                        labels.append(label)
                        for image_file in image_files:
                            if not (image_file.endswith('.jpg') or image_file.endswith('.png') or image_file.endswith('.jpeg')):
                                messagebox.showinfo('Warning', f'Sorry, found foreign files in dataset {dataset_path}!')
                                exit_loop = True
                                check_labels = False
                                break
                elif os.path.isfile(label_folder):
                    messagebox.showinfo('Warning', f'Sorry, found foreign files in dataset {dataset_path}!')
                    check_labels = False
                    break
            if check_labels:
                return {
                    'dataset_path': dataset_path,
                    'words': labels
                }
    return None

@eel.expose
def dataset_edit_add_take_button_event(dataset_path, new_word, dataset_type):
    global on_add_page
    new_word = new_word.title()
    if dataset_path:
        if new_word:
            dataset_data = dataset_selected_path_file_folder_text_event(dataset_path, dataset_type)
            if dataset_data:
                if not new_word in dataset_data['words']:
                    on_add_page = True
                    return new_word
                else:
                    messagebox.showinfo('Warning', 'Sorry, the new word already exists in the dataset!') 
        else:
            messagebox.showinfo('Warning', 'Please fill in the new word you want to add!') 
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def add_control_start_button_event(dataset_path, new_word, dataset_type):
    add_control_stop_button_event()
    global on_add_landmark, on_add_image
    if dataset_type == 'Landmark File':
        on_add_landmark = {
            'dataset_path': dataset_path,
            'new_word': new_word
        }
    elif dataset_type == 'Image Folder':
        on_add_image = {
            'dataset_path': dataset_path,
            'new_word': new_word
        }

@eel.expose
def add_control_stop_button_event():
    eel.update_add_progress(0, 0)()
    stop_take_landmark_image()

@eel.expose
def dataset_edit_delete_word_button_event(dataset_path, word, dataset_type):
    if dataset_path:
        if word:
            result = messagebox.askyesno('Confirmation', f'Are you sure you want to remove the word {word} from the dataset {dataset_path}?')
            if result:
                if dataset_type == 'Landmark File':
                    df = pd.read_csv(dataset_path)
                    df = df[df['class'] != word]
                    df.to_csv(dataset_path, index=False)
                elif dataset_type == 'Image Folder':
                    shutil.rmtree(f'{dataset_path}/{word}')
                eel.to_dataset_page()()
        else:
            messagebox.showinfo('Warning', 'Please select the words you want to delete!') 
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')

@eel.expose
def dataset_edit_redata_word_button_event(dataset_path, word):
    global on_redata_page
    word = word.title()
    if dataset_path:
        if word:
            on_redata_page = True
            return word
        else:
            messagebox.showinfo('Warning', 'Please select the words you want to redata!') 
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def redata_control_start_button_event(dataset_path, new_word, dataset_type):
    redata_control_stop_button_event()
    global on_redata_landmark, on_redata_image
    if dataset_type == 'Landmark File':
        on_redata_landmark = {
            'dataset_path': dataset_path,
            'new_word': new_word
        }
    elif dataset_type == 'Image Folder':
        on_redata_image = {
            'dataset_path': dataset_path,
            'new_word': new_word
        }

@eel.expose
def redata_control_stop_button_event():
    eel.update_redata_progress(0, 0)()
    stop_take_landmark_image()

@eel.expose
def dataset_edit_relabel_change_button_event(dataset_path, old_word, new_word, dataset_type):
    old_word = old_word.title()
    new_word = new_word.title()
    if dataset_path:
        if old_word:
            if new_word:
                dataset_data = dataset_selected_path_file_folder_text_event(dataset_path, dataset_type)
                if dataset_data:
                    if not new_word in dataset_data['words']:
                        if dataset_type == 'Landmark File':                            
                            df = pd.read_csv(dataset_path)
                            df['class'] = df['class'].replace(old_word, new_word)
                            df.to_csv(dataset_path, index=False)
                        elif dataset_type == 'Image Folder':
                            os.rename(f'{dataset_path}/{old_word}', f'{dataset_path}/{new_word}')
                            folder = os.listdir(f'{dataset_path}/{new_word}')
                            for index, image in enumerate(folder):
                                os.rename(f'{dataset_path}/{new_word}/{image}', f'{dataset_path}/{new_word}/{new_word}_{index+1}.jpg')
                        eel.to_dataset_page()()                        
                    else:
                        messagebox.showinfo('Warning', 'Sorry, the new word already exists in the dataset!') 
            else:
                messagebox.showinfo('Warning', 'Please fill in the new word!') 
        else:
            messagebox.showinfo('Warning', 'Please select the words you want to relabel!')             
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')        

# ============================================================================================================
# training
@eel.expose
def training_control_select_path_button_event():
    training_path = filedialog.askopenfilename()
    training_graph_event(training_path)

@eel.expose
def training_control_train_button_event(dataset_path):
    if dataset_path:
        return True
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_decision_tree_create_button_event(dataset_path, training_path, training_name, random_state, max_depth):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if random_state:
                    if int(random_state) > 0:
                        if max_depth:
                            if int(max_depth) >= 0:
                                training_path = training_path + '/' + training_name
                                get_training_path = decision_tree(dataset_path, training_path, random_state, max_depth)
                                return get_training_path
                            else:
                                messagebox.showinfo('Warning', 'Please fill the max depth in range (0, N)!')                        
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the max depth first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the random state with a positive integer!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the random state first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_random_forest_create_button_event(dataset_path, training_path, training_name, random_state, n_estimators):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if random_state:
                    if int(random_state) > 0:
                        if n_estimators:
                            if int(n_estimators) >= 100:
                                training_path = training_path + '/' + training_name
                                get_training_path = random_forest(dataset_path, training_path, random_state, n_estimators)
                                return get_training_path                                                      
                            else:
                                messagebox.showinfo('Warning', 'Please fill the n-estimators in range (100, N)!')                        
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the n-estimators first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the random state with a positive integer!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the random state first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_k_nearest_neighbors_create_button_event(dataset_path, training_path, training_name, random_state, n_neighbors):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if random_state:
                    if int(random_state) > 0:
                        if n_neighbors:
                            if int(n_neighbors) >= 5:
                                training_path = training_path + '/' + training_name
                                get_training_path = k_nearest_neighbors(dataset_path, training_path, random_state, n_neighbors)
                                return get_training_path                                                           
                            else:
                                messagebox.showinfo('Warning', 'Please fill the n-neighbors in range (5, N)!')                        
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the n-neighbors first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the random state with a positive integer!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the random state first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_convolutional_neural_network_create_button_event(dataset_path, training_path, training_name, shuffle, epochs):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if shuffle:
                    if int(shuffle) >= 300:
                        if epochs:
                            if int(epochs) >= 1:
                                training_path = training_path + '/' + training_name
                                get_training_path = convolutional_neural_network(dataset_path, training_path, shuffle, epochs)
                                return get_training_path
                            else:
                                messagebox.showinfo('Warning', 'Please fill the epochs in range (1, N)!')                        
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the epochs first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the shuffle with more than 300!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the shuffle first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_support_vector_machine_create_button_event(dataset_path, training_path, training_name, random_state, c, kernel):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if random_state:
                    if int(random_state) > 0:
                        if c:
                            if int(c) >= 1:
                                if kernel:
                                    training_path = training_path + '/' + training_name
                                    get_training_path = support_vector_machine(dataset_path, training_path, random_state, c, kernel)
                                    return get_training_path                                                                              
                                else:
                                    messagebox.showinfo('Warning', 'Please fill in the kernel!')                                
                            else:
                                messagebox.showinfo('Warning', 'Please fill the c in range (1, N)!')                        
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the c first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the random state with a positive integer!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the random state first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None

@eel.expose
def form_create_training_naive_bayes_create_button_event(dataset_path, training_path, training_name, random_state, var_smoothing, step):
    if dataset_path:
        if training_path:
            if training_name:
                training_name = training_name.split('.')[0] + '.pkl'
                if random_state:
                    if int(random_state) > 0:
                        if var_smoothing:
                            if re.match(r'^1e-\d+$', var_smoothing):
                                if step:
                                    if re.match(r'^1e-\d+$', step):                       
                                        training_path = training_path + '/' + training_name
                                        get_training_path = naive_bayes(dataset_path, training_path, random_state, var_smoothing, step)
                                        return get_training_path   
                                    else:
                                        messagebox.showinfo('Warning', 'Please fill the step with this format (1e-...)!') 
                                else:
                                    messagebox.showinfo('Warning', 'Please fill in the step first!')                                                                                
                            else:
                                messagebox.showinfo('Warning', 'Please fill the var smoothing in range (1e-9, N)!')   
                        else:
                            messagebox.showinfo('Warning', 'Please fill in the var smoothing first!')                        
                    else:
                        messagebox.showinfo('Warning', 'Please fill the random state with a positive integer!')
                else:
                    messagebox.showinfo('Warning', 'Please fill in the random state first!')
            else:
                messagebox.showinfo('Warning', 'Please fill in the training name first!')
        else:
            messagebox.showinfo('Warning', 'Please fill in the training path first!')
    else:
        messagebox.showinfo('Warning', 'Please select the dataset first!')
    return None


@eel.expose
def training_control_delete_button_event(training_path):
    if training_path:
        results = messagebox.askyesno('Confirmation', 'Are you sure you want to delete this training file?')
        if results:
            os.remove(training_path)
            return True
    else:
        messagebox.showinfo('Warning', 'Please select the training first!')
    return False

@eel.expose
def training_graph_event(training_path):
    if training_path:
        if os.path.splitext(training_path)[1].lower() == '.pkl':
            trained_model = None
            with open(training_path, 'rb') as file:
                trained_model = pickle.load(file)
            check_model = True
            try:
                trained_model['algorithm']
                trained_model['class_names']
                trained_model['classification_report']
                trained_model['confusion_matrix']
                trained_model['error_rate']
                trained_model['epoch_loss']
                trained_model['epoch_accuracy']
                trained_model['trained_model']

                trained_model['error_rate']['range']
                trained_model['error_rate']['train_error']
                trained_model['error_rate']['test_error']
                trained_model['epoch_loss']['epoch']
                trained_model['epoch_loss']['training_loss']
                trained_model['epoch_loss']['validation_loss']
                trained_model['epoch_accuracy']['epoch']
                trained_model['epoch_accuracy']['training_accuracy']
                trained_model['epoch_accuracy']['validation_accuracy']
            except:
                check_model = False
            if check_model:
                eel.training_selected_path_file_folder_text_event(training_path)()
                temporary_trained_data = trained_model
                training_algorithm = trained_model['algorithm']
                if trained_model['classification_report'] is not None:
                    labels = list(trained_model['classification_report'].keys())[:-3]
                    precision = [trained_model['classification_report'][label]['precision'] for label in labels]
                    recall = [trained_model['classification_report'][label]['recall'] for label in labels]
                    f1_score = [trained_model['classification_report'][label]['f1-score'] for label in labels]
                    x = range(len(labels))
                    width = 0.2
                    plt.figure(figsize=(8, 6))
                    plt.bar(x, precision, width=width, label='Precision')
                    plt.bar([i + width for i in x], recall, width=width, label='Recall')
                    plt.bar([i + 2 * width for i in x], f1_score, width=width, label='F1-score')
                    plt.xlabel('Class')
                    plt.ylabel('Scores')
                    plt.title(f'Classification Report {training_algorithm}')
                    plt.xticks([i + width for i in x], labels, rotation=45)
                    plt.legend()
                    plt.tight_layout()

                    image_data = io.BytesIO()
                    plt.savefig(image_data, format='jpg')
                    image_data.seek(0)                    
                    image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                    plt.close()
                    eel.update_training_graph(temporary_trained_data, image_base64, 'classification_report')()
                if trained_model['confusion_matrix'] is not None:
                    confusion_matrix_list = []
                    for row in temporary_trained_data['confusion_matrix']:
                        row_data = []
                        for column in row:
                            row_data.append(int(column))
                        confusion_matrix_list.append(row_data)
                    temporary_trained_data['confusion_matrix'] = confusion_matrix_list
                    plt.figure(figsize=(8, 6))
                    sns.heatmap(trained_model['confusion_matrix'], annot=True, fmt='d', cmap='Blues', cbar=False, xticklabels=trained_model['class_names'], yticklabels=trained_model['class_names'])
                    plt.xlabel('Predicted Labels')
                    plt.ylabel('True Labels')
                    plt.title(f'Confusion Matrix {training_algorithm}')
                    
                    image_data = io.BytesIO()
                    plt.savefig(image_data, format='jpg')
                    image_data.seek(0)                    
                    image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                    plt.close()
                    eel.update_training_graph(temporary_trained_data, image_base64, 'confusion_matrix')()
                if trained_model['error_rate']['range'] is not None:
                    temporary_trained_data['error_rate']['range'] = list(temporary_trained_data['error_rate']['range'])
                    if trained_model['algorithm'] == 'Random Forest':
                        plt.figure(figsize=(8, 6))
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['train_error'], label='Train Error', marker='o', linestyle='-', color='red')
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['test_error'], label='Test Error', marker='o', linestyle='-', color='blue')
                        plt.xlabel('N-Estimators (Range)')
                        plt.ylabel('Error Rate')
                        plt.title(f'Error Rate Vs N-Estimators {training_algorithm}')
                        plt.legend()

                        image_data = io.BytesIO()
                        plt.savefig(image_data, format='jpg')
                        image_data.seek(0)                    
                        image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                        plt.close()
                        eel.update_training_graph(temporary_trained_data, image_base64, 'error_rate')()
                    elif trained_model['algorithm'] == 'K-Nearest Neighbors':
                        plt.figure(figsize=(8, 6))
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['train_error'], label='Train Error', marker='o', linestyle='-', color='red')
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['test_error'], label='Test Error', marker='o', linestyle='-', color='blue')
                        plt.xlabel('N-Neighbors (Range)')
                        plt.ylabel('Error Rate')
                        plt.title(f'Error Rate Vs N-Neighbors {training_algorithm}')
                        plt.legend()

                        image_data = io.BytesIO()
                        plt.savefig(image_data, format='jpg')
                        image_data.seek(0)                    
                        image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                        plt.close()
                        eel.update_training_graph(temporary_trained_data, image_base64, 'error_rate')()
                    elif trained_model['algorithm'] == 'Decision Tree':
                        plt.figure(figsize=(8, 6))
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['train_error'], label='Train Error', marker='o', linestyle='-', color='red')
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['test_error'], label='Test Error', marker='o', linestyle='-', color='blue')
                        plt.xlabel('Max Depth (Range)')
                        plt.ylabel('Error Rate')
                        plt.title(f'Error Rate Vs Max Depth {training_algorithm}')
                        plt.legend()

                        image_data = io.BytesIO()
                        plt.savefig(image_data, format='jpg')
                        image_data.seek(0)                    
                        image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                        plt.close()
                        eel.update_training_graph(temporary_trained_data, image_base64, 'error_rate')()
                    elif trained_model['algorithm'] == 'Support Vector Machine':
                        plt.figure(figsize=(8, 6))
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['train_error'], label='Train Error', marker='o', linestyle='-', color='red')
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['test_error'], label='Test Error', marker='o', linestyle='-', color='blue')
                        plt.xlabel('C (Range)')
                        plt.ylabel('Error Rate')
                        plt.title(f'Error Rate Vs C {training_algorithm}')
                        plt.legend()

                        image_data = io.BytesIO()
                        plt.savefig(image_data, format='jpg')
                        image_data.seek(0)                    
                        image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                        plt.close()
                        eel.update_training_graph(temporary_trained_data, image_base64, 'error_rate')()
                    elif trained_model['algorithm'] == 'Naive Bayes':
                        plt.figure(figsize=(8, 6))
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['train_error'], label='Train Error', marker='o', linestyle='-', color='red')
                        plt.plot(trained_model['error_rate']['range'], trained_model['error_rate']['test_error'], label='Test Error', marker='o', linestyle='-', color='blue')
                        plt.xlabel('Var Smoothing (Range)')
                        plt.ylabel('Error Rate')
                        plt.title(f'Error Rate Vs Var Smoothing {training_algorithm}')
                        plt.legend()

                        image_data = io.BytesIO()
                        plt.savefig(image_data, format='jpg')
                        image_data.seek(0)                    
                        image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                        plt.close()
                        eel.update_training_graph(temporary_trained_data, image_base64, 'error_rate')()                                               
                if trained_model['epoch_loss']['epoch'] is not None:
                    temporary_trained_data['epoch_loss']['epoch'] = list(temporary_trained_data['epoch_loss']['epoch'])
                    plt.figure(figsize=(8, 6))
                    plt.plot(trained_model['epoch_loss']['epoch'], trained_model['epoch_loss']['training_loss'], label='Training Loss', marker='o', linestyle='-', color='red')
                    plt.plot(trained_model['epoch_loss']['epoch'], trained_model['epoch_loss']['validation_loss'], label='Validation Loss', marker='o', linestyle='-', color='blue')
                    plt.xlabel('Epochs')
                    plt.ylabel('Loss')
                    plt.title(f'Epochs Vs Loss {training_algorithm}')
                    plt.legend()

                    image_data = io.BytesIO()
                    plt.savefig(image_data, format='jpg')
                    image_data.seek(0)                    
                    image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                    plt.close()
                    eel.update_training_graph(temporary_trained_data, image_base64, 'epoch_loss')()
                if trained_model['epoch_accuracy']['epoch'] is not None:                    
                    temporary_trained_data['epoch_accuracy']['epoch'] = list(temporary_trained_data['epoch_accuracy']['epoch'])
                    plt.figure(figsize=(8, 6))
                    plt.plot(trained_model['epoch_accuracy']['epoch'], trained_model['epoch_accuracy']['training_accuracy'], label='Training Accuracy', marker='o', linestyle='-', color='red')
                    plt.plot(trained_model['epoch_accuracy']['epoch'], trained_model['epoch_accuracy']['validation_accuracy'], label='Validation Accuracy', marker='o', linestyle='-', color='blue')
                    plt.xlabel('Epochs')
                    plt.ylabel('Accuracy')
                    plt.title(f'Epochs Vs Accuracy {training_algorithm}')
                    plt.legend()

                    image_data = io.BytesIO()
                    plt.savefig(image_data, format='jpg')
                    image_data.seek(0)                    
                    image_base64 = base64.b64encode(image_data.read()).decode('utf-8')                    
                    plt.close()
                    eel.update_training_graph(temporary_trained_data, image_base64, 'epoch_accuracy')()
            else:
                messagebox.showinfo('Warning', 'Sorry, the training file is not compatible with our reader system!')            
        else:
            messagebox.showinfo('Warning', 'Sorry, the training file is not compatible with our reader system!')

def random_forest(dataset_path, training_path, random_state, n_estimators):
    random_state = int(random_state)
    n_estimators = int(n_estimators)
    df = pd.read_csv(dataset_path)
    X = df.drop('class', axis=1)
    y = df['class']
    key = {}.fromkeys(y)
    class_names = list(key.keys())
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=random_state)
    model = None
    train_error = []
    test_error = []
    oob_error = []
    n_estimators_range = range(100, n_estimators + 1)
    for trees in n_estimators_range:    
        rf = RandomForestClassifier(n_estimators=trees)
        rf.fit(X_train, y_train)
        train_error_rate = 1 - rf.score(X_train, y_train)
        test_error_rate = 1 - rf.score(X_test, y_test)
        # oob_error_rate = 1 - rf.oob_score_
        train_error.append(train_error_rate)
        test_error.append(test_error_rate)
        # oob_error.append(oob_error_rate)
        percentage = ((trees - 99) / (n_estimators - 99)) * 100
        eel.update_random_forest_progress(percentage, trees)()
        model = rf
    if model is not None:
        y_predict = model.predict(X_test)
        get_confusion_matrix = confusion_matrix(y_test, y_predict)
        get_classification_report = classification_report(y_test, y_predict, output_dict=True)
        model_training = {
            'algorithm': 'Random Forest',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': n_estimators_range,
                'train_error': train_error,
                'test_error': test_error
            },
            'epoch_loss': {
                'epoch': None,
                'training_loss': None,
                'validation_loss': None
            },
            'epoch_accuracy': {
                'epoch': None,
                'training_accuracy': None,
                'validation_accuracy': None
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

def decision_tree(dataset_path, training_path, random_state, max_depth):
    random_state = int(random_state)
    max_depth = int(max_depth)
    df = pd.read_csv(dataset_path)
    X = df.drop('class', axis=1)
    y = df['class']
    key = {}.fromkeys(y)
    class_names = list(key.keys())
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=random_state)
    model = None
    max_depth_range = None
    train_error = []
    test_error = []
    if max_depth == 0:
        max_depth_range = range(0, 1)
        dt = DecisionTreeClassifier(max_depth=None)
        dt.fit(X_train, y_train)
        train_error_rate = 1 - dt.score(X_train, y_train)
        test_error_rate = 1 - dt.score(X_test, y_test)
        train_error.append(train_error_rate)
        test_error.append(test_error_rate)
        percentage = 100
        eel.update_decision_tree_progress(percentage, 'none')()
        model = dt
    else:
        max_depth_range = range(1, max_depth + 1)
        for trees in max_depth_range:    
            dt = DecisionTreeClassifier(max_depth=trees)
            dt.fit(X_train, y_train)
            train_error_rate = 1 - dt.score(X_train, y_train)
            test_error_rate = 1 - dt.score(X_test, y_test)
            train_error.append(train_error_rate)
            test_error.append(test_error_rate)
            percentage = (trees / max_depth) * 100
            eel.update_decision_tree_progress(percentage, trees)()
            model = dt
    if model is not None:
        y_predict = model.predict(X_test)
        get_confusion_matrix = confusion_matrix(y_test, y_predict)
        get_classification_report = classification_report(y_test, y_predict, output_dict=True)
        model_training = {
            'algorithm': 'Decision Tree',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': max_depth_range,
                'train_error': train_error,
                'test_error': test_error
            },
            'epoch_loss': {
                'epoch': None,
                'training_loss': None,
                'validation_loss': None
            },
            'epoch_accuracy': {
                'epoch': None,
                'training_accuracy': None,
                'validation_accuracy': None
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

def k_nearest_neighbors(dataset_path, training_path, random_state, n_neighbors):
    random_state = int(random_state)
    n_neighbors = int(n_neighbors)
    df = pd.read_csv(dataset_path)
    X = df.drop('class', axis=1)
    y = df['class']
    key = {}.fromkeys(y)
    class_names = list(key.keys())
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=random_state)
    model = None
    train_error = []
    test_error = []
    n_neighbors_range = range(5, n_neighbors + 1)
    for trees in n_neighbors_range:    
        knn = KNeighborsClassifier(n_neighbors=trees)
        knn.fit(X_train, y_train)
        train_error_rate = 1 - knn.score(X_train, y_train)
        test_error_rate = 1 - knn.score(X_test, y_test)
        train_error.append(train_error_rate)
        test_error.append(test_error_rate)
        percentage = ((trees - 4) / (n_neighbors - 4)) * 100
        eel.update_k_nearest_neighbors_progress(percentage, trees)()
        model = knn
    if model is not None:
        y_predict = model.predict(X_test)
        get_confusion_matrix = confusion_matrix(y_test, y_predict)
        get_classification_report = classification_report(y_test, y_predict, output_dict=True)
        model_training = {
            'algorithm': 'K-Nearest Neighbors',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': n_neighbors_range,
                'train_error': train_error,
                'test_error': test_error
            },
            'epoch_loss': {
                'epoch': None,
                'training_loss': None,
                'validation_loss': None
            },
            'epoch_accuracy': {
                'epoch': None,
                'training_accuracy': None,
                'validation_accuracy': None
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

def convolutional_neural_network(dataset_path, training_path, shuffle, epochs):
    model = None
    shuffle = int(shuffle)
    epochs = int(epochs)
    data_directory = pathlib.Path(dataset_path)
    batch_size = 128
    image_height = 48
    image_width = 48
    training_dataset = tf.keras.utils.image_dataset_from_directory(
        data_directory,
        validation_split=0.2,
        subset='training',
        seed=123,
        image_size=(image_height, image_width),
        batch_size=batch_size,
        color_mode='grayscale'
    )
    validation_dataset = tf.keras.utils.image_dataset_from_directory(
        data_directory,
        validation_split=0.2,
        subset='validation',
        seed=123,
        image_size=(image_height, image_width),
        batch_size=batch_size,
        color_mode='grayscale'
    )
    class_names = training_dataset.class_names
    AUTOTUNE = tf.data.AUTOTUNE
    training_dataset = training_dataset.cache().shuffle(shuffle).prefetch(buffer_size=AUTOTUNE)
    validation_dataset = validation_dataset.cache().prefetch(buffer_size=AUTOTUNE)    
    normalization_layer = layers.Rescaling(1./255)
    training_dataset_normalized = training_dataset.map(lambda x, y: (normalization_layer(x), y))
    validation_dataset_normalized = validation_dataset.map(lambda x, y: (normalization_layer(x), y))
    number_classes = len(class_names)
    model = Sequential([
        layers.Conv2D(16, 3, padding='same', activation='relu', input_shape=(image_height, image_width, 1)),
        layers.MaxPooling2D(),
        layers.Conv2D(32, 3, padding='same', activation='relu'),
        layers.MaxPooling2D(),
        layers.Conv2D(64, 3, padding='same', activation='relu'),
        layers.MaxPooling2D(),
        layers.Flatten(),
        layers.Dense(128, activation='relu'),
        layers.Dense(number_classes)
    ])
    model.compile(
        optimizer='adam',
        loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
        metrics=['accuracy']
    )
    root.deiconify()
    root.geometry('1000x500+50+50')
    root.resizable(width=False, height=False)
    textbox = tk.Text(root, bg='black', fg='black', foreground='white', wrap='word')
    textbox.pack(side="left", fill="both", expand=True)
    scrollbar = tk.Scrollbar(root, command=textbox.yview)
    scrollbar.pack(side="right", fill="y")
    textbox.config(yscrollcommand=scrollbar.set)
    callback = tf.keras.callbacks.LambdaCallback(on_epoch_end=lambda epoch, logs: epoch_logger(textbox, epochs, epoch + 1, logs))
    history = model.fit(
        training_dataset_normalized,
        validation_data=validation_dataset_normalized,
        epochs=epochs,
        callbacks=[callback]
    )
    root.withdraw()
    if model is not None:
        training_accuracy = history.history['accuracy']
        validation_accuracy = history.history['val_accuracy']
        training_loss = history.history['loss']
        validation_loss = history.history['val_loss']
        epochs_range = range(1, epochs + 1)
        y_value_true = np.concatenate([y for x, y in validation_dataset], axis=0)
        y_value_predict = np.argmax(model.predict(validation_dataset_normalized), axis=1)
        get_confusion_matrix = confusion_matrix(y_value_true, y_value_predict)
        get_classification_report = classification_report(y_value_true, y_value_predict, target_names=class_names, output_dict=True)
        model_training = {
            'algorithm': 'Convolutional Neural Network',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': None,
                'train_error': None,
                'test_error': None
            },
            'epoch_loss': {
                'epoch': epochs_range,
                'training_loss': training_loss,
                'validation_loss': validation_loss
            },
            'epoch_accuracy': {
                'epoch': range(1, epochs + 1),
                'training_accuracy': training_accuracy,
                'validation_accuracy': validation_accuracy
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

def support_vector_machine(dataset_path, training_path, random_state, c, kernel):
    random_state = int(random_state)
    c = int(c)
    df = pd.read_csv(dataset_path)
    X = df.drop('class', axis=1)
    y = df['class']
    key = {}.fromkeys(y)
    class_names = list(key.keys())
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=random_state)
    model = None
    train_error = []
    test_error = []
    c_range = range(1, c + 1)
    for trees in c_range:    
        svm = SVC(C=trees, kernel=kernel, probability=True)
        svm.fit(X_train, y_train)
        train_error_rate = 1 - svm.score(X_train, y_train)
        test_error_rate = 1 - svm.score(X_test, y_test)
        train_error.append(train_error_rate)
        test_error.append(test_error_rate)
        percentage = (trees / c) * 100
        eel.update_support_vector_machine_progress(percentage, trees)()
        model = svm
    if model is not None:
        y_predict = model.predict(X_test)
        get_confusion_matrix = confusion_matrix(y_test, y_predict)
        get_classification_report = classification_report(y_test, y_predict, output_dict=True)
        model_training = {
            'algorithm': 'Support Vector Machine',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': c_range,
                'train_error': train_error,
                'test_error': test_error
            },
            'epoch_loss': {
                'epoch': None,
                'training_loss': None,
                'validation_loss': None
            },
            'epoch_accuracy': {
                'epoch': None,
                'training_accuracy': None,
                'validation_accuracy': None
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

def naive_bayes(dataset_path, training_path, random_state, var_smoothing, step):
    random_state = int(random_state)
    var_smoothing = float(var_smoothing)
    step = float(step)
    df = pd.read_csv(dataset_path)
    X = df.drop('class', axis=1)
    y = df['class']
    key = {}.fromkeys(y)
    class_names = list(key.keys())
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=random_state)
    model = None
    train_error = []
    test_error = []
    var_smoothing_range = np.arange(1e-9, var_smoothing, step)
    for trees in var_smoothing_range:    
        nb = GaussianNB(var_smoothing=trees)
        nb.fit(X_train, y_train)
        train_error_rate = 1 - nb.score(X_train, y_train)
        test_error_rate = 1 - nb.score(X_test, y_test)
        train_error.append(train_error_rate)
        test_error.append(test_error_rate)
        percentage = (trees / var_smoothing) * 100
        eel.update_naive_bayes_progress(percentage, trees)()
        model = nb
    if model is not None:
        y_predict = model.predict(X_test)
        get_confusion_matrix = confusion_matrix(y_test, y_predict)
        get_classification_report = classification_report(y_test, y_predict, output_dict=True)
        model_training = {
            'algorithm': 'Naive Bayes',
            'class_names': class_names,
            'classification_report': get_classification_report,
            'confusion_matrix': get_confusion_matrix,
            'error_rate': {
                'range': var_smoothing_range,
                'train_error': train_error,
                'test_error': test_error
            },
            'epoch_loss': {
                'epoch': None,
                'training_loss': None,
                'validation_loss': None
            },
            'epoch_accuracy': {
                'epoch': None,
                'training_accuracy': None,
                'validation_accuracy': None
            },
            'trained_model': model
        }
        with open(training_path, 'wb') as file:
            pickle.dump(model_training, file)
        return training_path
    return None

@eel.expose
def azure_machine_learning_cnn():
    messagebox.showinfo('Warning', 'Sorry, CNN training with Azure Machine Learning is still under development!')            

def epoch_logger(textbox, epochs, epoch, logs):
    root.update_idletasks()
    new_text = f'''Epoch {epoch + 1}/{epochs}:
loss: {logs["loss"]:.4f} - accuracy: {logs["accuracy"]:.4f} - val_loss: {logs["val_loss"]:.4f} - val_accuracy: {logs["val_accuracy"]:.4f}
'''
    textbox.insert(tk.END, new_text)
    textbox.see(tk.END)

# ============================================================================================================
# translate
    
@eel.expose
def navigator_to_translate_page_event():
    global on_translate_page
    on_translate_page = True

@eel.expose
def navigator_to_other_than_translate_page_event():
    global on_translate_page
    on_translate_page = False

@eel.expose
def translate_control_marker_button_event():
    global translate_marker
    if translate_marker:
        translate_marker = False
    else:
        translate_marker = True
    return translate_marker

@eel.expose
def translate_control_start_button_event(training_path, history_path):
    global on_tranlsate_with_training_model, with_history_path
    if on_tranlsate_with_training_model is None:
        if training_path:
            if history_path:
                trained_model = None
                with open(training_path, 'rb') as file:
                    trained_model = pickle.load(file)
                on_tranlsate_with_training_model = trained_model
                with_history_path = history_path
            else:
                results = messagebox.askyesno('Confirmation', 'You haven\'t selected a history file, do you want to continue translating without a history file?')
                if results:
                    trained_model = None
                    with open(training_path, 'rb') as file:
                        trained_model = pickle.load(file)
                    on_tranlsate_with_training_model = trained_model
                    with_history_path = None
        else:
            messagebox.showinfo('Warning', 'Sorry, can\'t translate without training file!')

@eel.expose
def empty_check_translate():
    global check_translate, count_time
    check_translate = []
    count_time = 0

@eel.expose
def translate_control_stop_button_event():
    global on_tranlsate_with_training_model, count_time, history_data, with_history_path, check_translate
    update_history = None
    if on_tranlsate_with_training_model is not None:
        if with_history_path is not None:
            if len(history_data) > 0:            
                workbook = load_workbook(with_history_path)                
                sheet_name = history_data[0]['date'] + ' ' + history_data[0]['time'].replace(':', '_')
                sheet = workbook.create_sheet(title=sheet_name)            
                header = ['Translation', 'Date', 'Time']
                sheet.append(header)
                bold_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for cell in sheet[1]:
                    cell.font = bold_font
                    cell.border = thin_border
                for item in history_data:
                    row = [item['translate'], item['date'], item['time']]
                    sheet.append(row)
                workbook.save(with_history_path)
                update_history = with_history_path
    eel.send_text_translate_control_result_text_event('')()
    eel.translate_control_time_limit_progressbar_event(0)()
    eel.clear_translate_translation_list_sub_frame_event()()
    check_translate = []
    history_data = []
    on_tranlsate_with_training_model = None
    count_time = 0
    with_history_path = None
    return update_history

# ============================================================================================================
# history

@eel.expose
def history_control_select_path_button_event():
    history_path = filedialog.askopenfilename()
    return history_path

@eel.expose
def form_history_control_select_path_button_event():
    history_path = filedialog.askdirectory()
    return history_path

@eel.expose
def history_selected_path_file_folder_text_event(history_path):
    if history_path:
        if os.path.splitext(history_path)[1].lower() == '.xlsx':
            data = pd.read_excel(history_path, sheet_name=None)
            get_data = []        
            for sheet_name, sheet_data in data.items():
                if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}$', sheet_name):
                    temp_sheet = sheet_name.split(' ')
                    temp_sheet[1] = temp_sheet[1].replace('_', ':')
                    sheet_id = temp_sheet[0] + ' | ' + temp_sheet[1]
                    get_data.append({
                        'sheet': sheet_id,
                        'values': sheet_data.values.tolist()
                    })
            return get_data
        else:
            messagebox.showinfo('Warning', 'Sorry, the history file is not compatible with our reader system!')
    return None

@eel.expose
def history_selected_date_time_delete_button_event(history_path, sheet_name):
    if history_path:
        if sheet_name:
            workbook = openpyxl.load_workbook(history_path)
            sheet_names = workbook.sheetnames
            if len(sheet_names) == 1:
                results = messagebox.askyesno('Confirmation', 'You only have one sheet left, if you still want to delete, the history file will be deleted, are you sure you want to delete?')
                if results:
                    os.remove(history_path)
                    return 'delete_file'
            else:
                results = messagebox.askyesno('Confirmation', 'Are you sure you want to delete this translation?')
                if results:
                    sheet_to_remove = workbook[sheet_name]
                    workbook.remove(sheet_to_remove)
                    workbook.save(history_path)
                    return 'delete_sheet'
        else:
            messagebox.showinfo('Warning', 'Please select the date time of translation list first!')
    else:
        messagebox.showinfo('Warning', 'Please select the history first!')
    return None

@eel.expose
def history_control_delete_button_event(history_path):
    if history_path:
        results = messagebox.askyesno('Confirmation', 'Are you sure you want to delete this history file?')
        if results:
            os.remove(history_path)
            return True
    else:
        messagebox.showinfo('Warning', 'Please select the history first!')
    return False

@eel.expose
def history_control_create_button_event(history_path, history_name):
    if history_path:
        history_path = history_path + '/' + history_name.split('.')[0] + '.xlsx'
        if history_name:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Signfinity'
            sheet['A1'] = 'Welcome, you can save your signfinity sign language translation history here :)'
            workbook.save(history_path)
            return history_path
        else:
            messagebox.showinfo('Warning', 'Please complete the history name section of the form!')
    else:
        messagebox.showinfo('Warning', 'Please complete the history path section of the form!')
    return None

# ============================================================================================================

eel.start('index.html')